import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from database.db import get_all_specialists_full


async def export_specialists_to_excel(filepath: str):
    """Barcha mutaxassislarni Excel faylga eksport qiladi"""
    specialists = await get_all_specialists_full()

    wb = Workbook()
    ws = wb.active
    ws.title = "Mutaxassislar"

    headers = ["ID", "Ism", "Kategoriya", "Kasbi", "Telefon", "Tavsif",
               "Tavsiyalar", "Reyting", "Holat", "Qo'shilgan sana"]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for s in specialists:
        rating_count = s.get('rating_count', 0) or 0
        rating_sum = s.get('rating_sum', 0) or 0
        avg_rating = round(rating_sum / rating_count, 1) if rating_count > 0 else "-"
        status = "✅ Tasdiqlangan" if s['approved'] else "⏳ Kutilmoqda"

        ws.append([
            s['id'],
            s['fullname'],
            s.get('category', '-') or '-',
            s['profession'],
            s['phone'],
            s.get('description', '') or '',
            s['recommendation_count'],
            avg_rating,
            status,
            s.get('created_at', ''),
        ])

    # Ustun kengligi
    widths = [6, 22, 28, 22, 16, 40, 12, 10, 16, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    return filepath
